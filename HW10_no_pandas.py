"""Text Topic Analysis on Latin American Wine Reviews (no pandas/numpy).

This script reads review data from an Excel workbook, cleans and normalises the
text, extracts TF–IDF features, and fits an LDA topic model.  It also prints a
summary showing the number of reviews, average points, and average price for
each discovered topic.  Optional word clouds are generated when the
`wordcloud` package is available.

Improvements over the original version:
  * Robust Excel loading (skips headers correctly and validates columns)
  * Automatic NLTK resource checks/downloads to avoid runtime lookup errors
  * Safer token normalisation and preprocessing pipeline
  * Clear error messages when required resources or files are missing
"""

from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path
from typing import Dict, Iterable, List, Sequence, Tuple

import matplotlib.pyplot as plt
import nltk
from nltk import pos_tag, word_tokenize
from nltk.corpus import stopwords
from nltk.stem import PorterStemmer
from openpyxl import load_workbook
from sklearn.decomposition import LatentDirichletAllocation
from sklearn.feature_extraction.text import TfidfVectorizer


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

DATA_FILE = Path("LatinAmericanWines.xlsx")

# Domain specific stop words and synonym replacements
DOMAIN_STOPWORDS = {
    "wine",
    "wines",
    "flavor",
    "flavors",
    "aroma",
    "aromas",
    "palate",
    "drink",
}

SYNONYM_MAP = {
    "blackberries": "berry",
    "blackberry": "berry",
    "berries": "berry",
    "cherries": "cherry",
    "raspberries": "raspberry",
    "citrusy": "citrus",
    "lime": "citrus",
    "lemony": "citrus",
    "grapefruit": "citrus",
    "vanillin": "vanilla",
    "oaky": "oak",
}

N_COMPONENTS = 5
TOP_TERMS_PER_TOPIC = 15


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def ensure_nltk_resource(resource_path: str, download_name: str) -> None:
    """Ensure the given NLTK resource is available, download if missing."""

    try:
        nltk.data.find(resource_path)
    except LookupError:
        nltk.download(download_name)


def ensure_nltk_dependencies() -> None:
    """Download required NLTK resources when they are not present."""

    ensure_nltk_resource("tokenizers/punkt", "punkt")
    # Averaged perceptron tagger changed name in newer NLTK releases; try both.
    try:
        ensure_nltk_resource("taggers/averaged_perceptron_tagger", "averaged_perceptron_tagger")
    except LookupError:
        ensure_nltk_resource(
            "taggers/averaged_perceptron_tagger_eng",
            "averaged_perceptron_tagger_eng",
        )
    ensure_nltk_resource("corpora/stopwords", "stopwords")


def load_excel_data(path: Path) -> Tuple[List[str], List[float], List[float]]:
    """Load descriptions, points, and prices from the Excel workbook."""

    if not path.is_file():
        raise FileNotFoundError(
            f"Expected data file '{path.name}' in {path.parent.resolve()}"
        )

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        rows = wb.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            return [], [], []

        header_map = {
            (value.strip() if isinstance(value, str) else value): idx
            for idx, value in enumerate(header)
            if value is not None
        }

        required_columns = {"Description", "Points", "Price"}
        missing = required_columns.difference(header_map)
        if missing:
            raise ValueError(
                f"Missing expected columns in Excel file: {', '.join(sorted(missing))}"
            )

        desc_idx = header_map["Description"]
        points_idx = header_map["Points"]
        price_idx = header_map["Price"]

        descriptions: List[str] = []
        points: List[float] = []
        prices: List[float] = []

        for row in rows:
            if row is None:
                continue

            desc_value = row[desc_idx] if len(row) > desc_idx else None
            point_value = row[points_idx] if len(row) > points_idx else None
            price_value = row[price_idx] if len(row) > price_idx else None

            descriptions.append(str(desc_value) if desc_value is not None else "")

            try:
                points.append(float(point_value) if point_value is not None else None)
            except (TypeError, ValueError):
                points.append(None)

            try:
                prices.append(float(price_value) if price_value is not None else None)
            except (TypeError, ValueError):
                prices.append(None)

        return descriptions, points, prices
    finally:
        wb.close()


def build_stopwords() -> set:
    """Combine standard English stopwords with domain-specific ones."""

    eng_stop = set(stopwords.words("english"))
    return eng_stop.union(DOMAIN_STOPWORDS)


def normalise_token(token: str) -> str:
    token = token.lower()
    token = re.sub(r"[^a-z]+", "", token)
    if not token:
        return ""
    return SYNONYM_MAP.get(token, token)


def preprocess_documents(texts: Sequence[str], stop_words: set, stemmer: PorterStemmer) -> List[str]:
    processed: List[str] = []
    for text in texts:
        if not text:
            processed.append("")
            continue

        tokens = word_tokenize(text)
        tagged = pos_tag(tokens)
        kept_tokens: List[str] = []

        for word, tag in tagged:
            norm = normalise_token(word)
            if len(norm) <= 2 or not norm or norm in stop_words:
                continue
            if not (tag.startswith("NN") or tag.startswith("JJ")):
                continue
            stemmed = stemmer.stem(norm)
            if stemmed:
                kept_tokens.append(stemmed)

        processed.append(" ".join(kept_tokens))

    return processed


def top_terms(model: LatentDirichletAllocation, features: Sequence[str], n: int) -> List[Tuple[int, List[str]]]:
    topics: List[Tuple[int, List[str]]] = []
    for idx, component in enumerate(model.components_):
        top_indices = component.argsort()[::-1][:n]
        topics.append((idx, [features[j] for j in top_indices]))
    return topics


def describe_topics(
    assignments: Sequence[int],
    topic_matrix: Sequence[Sequence[float]],
    points: Sequence[float],
    prices: Sequence[float],
) -> Dict[int, Dict[str, float]]:
    summary: Dict[int, Dict[str, float]] = defaultdict(
        lambda: {"n": 0, "pts": 0.0, "prc": 0.0, "count_pts": 0, "count_prc": 0}
    )

    for idx, topic in enumerate(assignments):
        topic_data = summary[topic]
        topic_data["n"] += 1

        point_val = points[idx]
        if point_val is not None:
            topic_data["pts"] += point_val
            topic_data["count_pts"] += 1

        price_val = prices[idx]
        if price_val is not None:
            topic_data["prc"] += price_val
            topic_data["count_prc"] += 1

    return summary


def print_topic_summary(summary: Dict[int, Dict[str, float]]) -> None:
    print("\nSummary by Topic:")
    print("Topic | #Reviews | AvgPoints | AvgPrice")
    for topic in sorted(summary):
        stats = summary[topic]
        avg_pts = stats["pts"] / stats["count_pts"] if stats["count_pts"] else 0.0
        avg_prc = stats["prc"] / stats["count_prc"] if stats["count_prc"] else 0.0
        print(f"{topic:5d} | {stats['n']:8d} | {avg_pts:9.2f} | {avg_prc:8.2f}")


def generate_wordclouds(
    lda_model: LatentDirichletAllocation,
    topics: Sequence[Tuple[int, Sequence[str]]],
    feature_names: Sequence[str],
) -> None:
    try:
        from wordcloud import WordCloud
    except ImportError:
        print("WordCloud not installed; skipping clouds.")
        return

    for topic_idx, _ in topics:
        weights = lda_model.components_[topic_idx]
        top_indices = weights.argsort()[::-1][:TOP_TERMS_PER_TOPIC]
        frequencies = {feature_names[j]: float(weights[j]) for j in top_indices}

        wordcloud = WordCloud(width=800, height=400, background_color="white")
        wordcloud = wordcloud.generate_from_frequencies(frequencies)

        plt.figure(figsize=(5, 3))
        plt.imshow(wordcloud, interpolation="bilinear")
        plt.axis("off")
        plt.title(f"Topic {topic_idx}")
        plt.tight_layout()
        plt.show()


def main() -> None:
    ensure_nltk_dependencies()

    descriptions, points, prices = load_excel_data(DATA_FILE)
    if not descriptions:
        print("No data rows found in the Excel file.")
        return

    print("Preprocessing text ... this may take a few minutes")
    stop_words = build_stopwords()
    stemmer = PorterStemmer()
    processed_docs = preprocess_documents(descriptions, stop_words, stemmer)

    vectorizer = TfidfVectorizer(max_df=0.95, min_df=0.01)
    matrix = vectorizer.fit_transform(processed_docs)

    lda_model = LatentDirichletAllocation(n_components=N_COMPONENTS, random_state=42)
    doc_topic = lda_model.fit_transform(matrix)
    feature_names = vectorizer.get_feature_names_out()

    topics = top_terms(lda_model, feature_names, TOP_TERMS_PER_TOPIC)
    for idx, terms in topics:
        print(f"\nTopic {idx}: {', '.join(terms)}")

    assignments = doc_topic.argmax(axis=1)
    summary = describe_topics(assignments, doc_topic, points, prices)
    print_topic_summary(summary)

    generate_wordclouds(lda_model, topics, feature_names)


if __name__ == "__main__":
    main()
